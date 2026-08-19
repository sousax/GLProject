# -*- coding: utf-8 -*-
"""
gl_report.py
-------------
Gera a planilha final consolidada (no espírito do "GL_TESTE.xlsx" usado
como modelo): uma aba "GL" com o resumo pronto para envio (faturas, PO/Item,
valor total somado, incoterm, peso bruto) e uma aba "Validacao" com o
resultado do cross-check, para auditoria antes do envio.
"""

from typing import List, Dict, Optional
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from invoice_extractor import InvoiceData
from packing_list_extractor import POItemSummary
from cross_check import Finding


def _fmt_item(item: str) -> str:
    """'000010' -> '10' (formato usual de item de PO sem zeros à esquerda)."""
    if item and item.isdigit():
        return str(int(item))
    return item


def _po_items_for_invoice(inv: InvoiceData, pl_summaries: Optional[Dict[str, POItemSummary]]) -> list:
    """
    Descobre o(s) número(s) de item da PO para uma fatura, priorizando:
      1. 'po_line' extraído diretamente da própria fatura (mais confiável
         quando existe, pois é isso que o fornecedor está usando de fato);
      2. senão, o 'PO Item' da packing list (fallback de quando a fatura
         não expõe esse número).
    """
    po_lines = sorted({_fmt_item(li.po_line) for li in inv.line_items if li.po_line})
    if po_lines:
        return po_lines
    if pl_summaries:
        return sorted({_fmt_item(s.po_item) for s in pl_summaries.values() if s.po == inv.customer_po})
    return []


def _resolve_ncm(invoices: List[InvoiceData], ncm_arg: str) -> str:
    """NCM final do relatório: usa o que foi passado manualmente (--ncm),
    e completa com os NCMs detectados automaticamente nas faturas."""
    detected = sorted({inv.ncm for inv in invoices if inv.ncm})
    if ncm_arg and detected:
        extra = [d for d in detected if d != ncm_arg]
        return ncm_arg + (f" (detectado nas faturas: {', '.join(extra)})" if extra else "")
    if ncm_arg:
        return ncm_arg
    if detected:
        return " / ".join(detected)
    return ""

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
LABEL_FONT = Font(name="Arial", bold=True)
NORMAL_FONT = Font(name="Arial")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

LEVEL_FILL = {
    "OK": PatternFill("solid", fgColor="C6E0B4"),
    "DIVERGENCIA": PatternFill("solid", fgColor="F4C7C3"),
    "AVISO": PatternFill("solid", fgColor="FFE699"),
}


def _set(ws, cell, value, font=NORMAL_FONT, align=None):
    ws[cell] = value
    ws[cell].font = font
    if align:
        ws[cell].alignment = align


def build_gl_workbook(
    invoices: List[InvoiceData],
    pl_summaries: Optional[Dict[str, POItemSummary]] = None,
    findings: Optional[List[Finding]] = None,
    importador: str = "",
    cnpj: str = "",
    modal: str = "",
    ncm: str = "",
) -> Workbook:
    wb = Workbook()

    # ------------------------------------------------------------------ GL
    ws = wb.active
    ws.title = "GL"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 20

    _set(ws, "A1", "Resumo para GL", Font(name="Arial", bold=True, size=14))
    ws.merge_cells("A1:D1")

    rows = [
        ("Importador", importador),
        ("CNPJ", cnpj),
        ("Modal", modal),
        ("NCM", _resolve_ncm(invoices, ncm)),
    ]
    r = 3
    for label, val in rows:
        _set(ws, f"A{r}", label, LABEL_FONT)
        _set(ws, f"B{r}", val)
        r += 1

    incoterms = sorted({inv.incoterm for inv in invoices if inv.incoterm})
    _set(ws, f"A{r}", "Incoterm", LABEL_FONT)
    _set(ws, f"B{r}", " / ".join(incoterms) if incoterms else "")
    r += 1

    currencies = sorted({inv.currency for inv in invoices if inv.currency})
    currency_label = currencies[0] if len(currencies) == 1 else " / ".join(currencies)
    r += 1

    # --- tabela de faturas -------------------------------------------------
    table_header_row = r
    headers = ["Fatura (Invoice)", "Data", "PO Cliente", "Item(ns) da PO",
               "Incoterm", "Moeda", "Valor (sem soma)"]
    for i, h in enumerate(headers):
        c = ws.cell(row=table_header_row, column=i + 1, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    first_data_row = table_header_row + 1
    row_cursor = first_data_row
    value_cells = []

    for inv in invoices:
        po_items = ""
        items = _po_items_for_invoice(inv, pl_summaries)
        if items:
            po_items = ", ".join(items)
        vals = [
            inv.invoice_number or "",
            inv.invoice_date or "",
            inv.customer_po or "",
            po_items,
            inv.incoterm or "",
            inv.currency or "",
            inv.total_value if inv.total_value is not None else "",
        ]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=row_cursor, column=col, value=v)
            c.font = NORMAL_FONT
            c.border = BORDER
            if col == 7 and isinstance(v, (int, float)):
                c.number_format = "#,##0.00"
        value_cells.append(f"G{row_cursor}")
        row_cursor += 1

    last_data_row = row_cursor - 1

    # linha de total (formula, nao valor fixo -> recalcula se editar a tabela)
    total_row = row_cursor + 1
    _set(ws, f"A{total_row}", "PO/Item (consolidado)", LABEL_FONT)
    po_item_consolidated = " / ".join(
        f"{inv.customer_po}-" + ",".join(_po_items_for_invoice(inv, pl_summaries))
        for inv in invoices if inv.customer_po
    )
    ws.merge_cells(f"B{total_row}:D{total_row}")
    _set(ws, f"B{total_row}", po_item_consolidated)

    total_row2 = total_row + 1
    _set(ws, f"A{total_row2}", "Valor total (soma das faturas)", LABEL_FONT)
    if last_data_row >= first_data_row:
        ws[f"C{total_row2}"] = f"=SUM(G{first_data_row}:G{last_data_row})"
    else:
        ws[f"C{total_row2}"] = 0
    ws[f"C{total_row2}"].number_format = "#,##0.00"
    ws[f"C{total_row2}"].font = Font(name="Arial", bold=True)
    _set(ws, f"D{total_row2}", currency_label, LABEL_FONT)

    if pl_summaries:
        total_gw_row = total_row2 + 1
        total_gw = sum(s.total_gross_weight for s in pl_summaries.values())
        total_nw = sum(s.total_net_weight for s in pl_summaries.values())
        _set(ws, f"A{total_gw_row}", "Peso Bruto total (packing list)", LABEL_FONT)
        _set(ws, f"B{total_gw_row}", f"{total_gw:.2f} KGS")
        _set(ws, f"A{total_gw_row+1}", "Peso Líquido total (packing list)", LABEL_FONT)
        _set(ws, f"B{total_gw_row+1}", f"{total_nw:.2f} KGS")

    _set(ws, f"A{ws.max_row+2}", f"Gerado automaticamente em {datetime.now():%Y-%m-%d %H:%M}",
         Font(name="Arial", italic=True, size=9, color="808080"))

    # ------------------------------------------------------------ Validacao
    wsv = wb.create_sheet("Validacao")
    wsv.sheet_view.showGridLines = False
    wsv.column_dimensions["A"].width = 14
    wsv.column_dimensions["B"].width = 24
    wsv.column_dimensions["C"].width = 90

    hdrs = ["Status", "Fatura", "Detalhe"]
    for i, h in enumerate(hdrs):
        c = wsv.cell(row=1, column=i + 1, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER

    r = 2
    for f in (findings or []):
        c0 = wsv.cell(row=r, column=1, value=f.level)
        c1 = wsv.cell(row=r, column=2, value=f.invoice)
        c2 = wsv.cell(row=r, column=3, value=f.message)
        for c in (c0, c1, c2):
            c.border = BORDER
            c.font = NORMAL_FONT
            c.alignment = Alignment(wrap_text=True, vertical="top")
        fill = LEVEL_FILL.get(f.level)
        if fill:
            c0.fill = fill
        r += 1

    # --------------------------------------------------------- Itens (raw)
    wsi = wb.create_sheet("Itens_Fatura")
    wsi.sheet_view.showGridLines = False
    hdrs = ["Fatura", "Ref Line", "Part Number", "Qtd", "Unidade",
            "Preço Unit.", "Valor s/ VAT", "Valor c/ VAT", "Ref PO (linha)",
            "PO Line (item da PO)", "NCM/HS"]
    for i, h in enumerate(hdrs):
        c = wsi.cell(row=1, column=i + 1, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
    r = 2
    for inv in invoices:
        for li in inv.line_items:
            vals = [inv.invoice_number, li.ref_line, li.part_number, li.qty, li.unit,
                    li.unit_price, li.amount_excl_vat, li.amount_incl_vat, li.ref_po,
                    li.po_line, li.hs_code]
            for col, v in enumerate(vals, start=1):
                c = wsi.cell(row=r, column=col, value=v)
                c.border = BORDER
                c.font = NORMAL_FONT
            r += 1
    for col in range(1, 12):
        wsi.column_dimensions[get_column_letter(col)].width = 16

    return wb
