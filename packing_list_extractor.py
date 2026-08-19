# -*- coding: utf-8 -*-
"""
packing_list_extractor.py
--------------------------
Lê a aba de Packing List (PL) de uma planilha Excel e devolve, por
PO + item, os dados agregados: quantidade de volumes, peso líquido,
peso bruto e a lista de números de série (SN), quando existir.

O formato de referência é a planilha "BRELE260728SEA.xlsx" (aba "PL"),
mas a leitura é feita procurando dinamicamente a linha de cabeçalho
(por nomes de coluna conhecidos), então funciona mesmo se a aba não
começar sempre na mesma linha.
"""

import re
from dataclasses import dataclass, field
from typing import Dict, List, Optional

import openpyxl

# Nomes de coluna esperados no cabeçalho da packing list (case-insensitive,
# comparação por "contém"). Ajuste/adicione sinônimos aqui se sua PL usar
# nomes diferentes.
COLUMN_ALIASES = {
    "pko_no": ["pko no"],
    "po": ["ilx po", "po no", "customer po", "po"],
    "po_item": ["po item", "item no"],
    "part_number": ["part number", "part no"],
    "description": ["description"],
    "qty": ["qty(pcs)", "qty"],
    "net_weight": ["n.w(kgs)", "net weight"],
    "gross_weight": ["g.w(kgs)", "gross weight"],
    "sn_no": ["sn no", "serial"],
    "hs_code": ["hs code", "ncm"],
    "coo": ["coo", "origin"],
}


@dataclass
class POItemSummary:
    po: str
    po_item: str
    part_numbers: List[str] = field(default_factory=list)
    total_qty: float = 0.0
    total_net_weight: float = 0.0
    total_gross_weight: float = 0.0
    n_packages: int = 0
    serial_numbers: List[str] = field(default_factory=list)
    hs_codes: List[str] = field(default_factory=list)
    countries_of_origin: List[str] = field(default_factory=list)


def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip().lower()


def _find_header_row(ws, max_scan_rows: int = 40):
    """Procura a linha de cabeçalho localizando pelo menos 3 colunas
    conhecidas (ex.: 'pko no', 'qty', 'n.w(kgs)') na mesma linha."""
    for r in range(1, min(max_scan_rows, ws.max_row) + 1):
        row_vals = [_norm(c.value) for c in ws[r]]
        hits = 0
        for aliases in COLUMN_ALIASES.values():
            if any(any(alias in v for v in row_vals) for alias in aliases):
                hits += 1
        if hits >= 4:
            return r
    return None


def _map_columns(ws, header_row: int) -> Dict[str, int]:
    row_vals = {c.column: _norm(c.value) for c in ws[header_row]}
    col_map = {}
    used_cols = set()
    for field_name, aliases in COLUMN_ALIASES.items():
        # Percorre os aliases EM ORDEM DE PRIORIDADE (o primeiro que casar
        # em alguma coluna vence) — evita que um alias genérico de outro
        # campo (ex.: "item no") roube a coluna de um alias mais específico
        # (ex.: "po item") só porque foi varrida antes.
        found = None
        for alias in aliases:
            for col_idx, val in row_vals.items():
                if col_idx in used_cols:
                    continue
                if alias in val:
                    found = col_idx
                    break
            if found is not None:
                break
        if found is not None:
            col_map[field_name] = found
            used_cols.add(found)
    return col_map


def extract_packing_list(xlsx_path: str, sheet_name: str = "PL") -> Dict[str, POItemSummary]:
    """
    Retorna um dicionário {"PO-ITEM": POItemSummary} agregando todas as
    linhas (volumes/caixas) da packing list por PO + item.

    Se `sheet_name` não existir no arquivo, tenta detectar automaticamente
    a aba que contém as colunas de packing list.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    ws = None
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        for sn in wb.sheetnames:
            if _find_header_row(wb[sn]) is not None:
                ws = wb[sn]
                break
    if ws is None:
        raise ValueError(
            f"Não foi possível localizar uma aba de packing list em {xlsx_path}"
        )

    header_row = _find_header_row(ws)
    if header_row is None:
        raise ValueError(f"Cabeçalho da packing list não encontrado na aba '{ws.title}'")

    col_map = _map_columns(ws, header_row)
    required = ["po", "po_item"]
    missing = [f for f in required if f not in col_map]
    if missing:
        raise ValueError(f"Colunas obrigatórias não encontradas na PL: {missing}")

    summaries: Dict[str, POItemSummary] = {}

    for r in range(header_row + 1, ws.max_row + 1):
        def get(field_name):
            col = col_map.get(field_name)
            if col is None:
                return None
            return ws.cell(row=r, column=col).value

        po = get("po")
        po_item = get("po_item")
        if po is None or po_item is None:
            continue
        po = str(po).strip()
        po_item = str(po_item).strip()
        if not po or not po_item or not po[0].isdigit():
            continue

        key = f"{po}-{po_item}"
        if key not in summaries:
            summaries[key] = POItemSummary(po=po, po_item=po_item)
        s = summaries[key]

        part_number = get("part_number")
        if part_number and part_number not in s.part_numbers:
            s.part_numbers.append(str(part_number))

        qty = get("qty")
        if isinstance(qty, (int, float)):
            s.total_qty += qty

        nw = get("net_weight")
        if isinstance(nw, (int, float)):
            s.total_net_weight += nw

        gw = get("gross_weight")
        if isinstance(gw, (int, float)):
            s.total_gross_weight += gw

        s.n_packages += 1

        sn = get("sn_no")
        if sn:
            s.serial_numbers.append(str(sn).strip())

        hs = get("hs_code")
        if hs and str(hs) not in s.hs_codes:
            s.hs_codes.append(str(hs))

        coo = get("coo")
        if coo and str(coo) not in s.countries_of_origin:
            s.countries_of_origin.append(str(coo))

    return summaries


def summaries_by_po(summaries: Dict[str, POItemSummary]) -> Dict[str, List[POItemSummary]]:
    """Agrupa os resumos por número de PO (útil para cross-check por PO)."""
    grouped: Dict[str, List[POItemSummary]] = {}
    for s in summaries.values():
        grouped.setdefault(s.po, []).append(s)
    return grouped
